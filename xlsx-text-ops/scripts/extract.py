"""Extract all text from an Excel (.xlsx/.xlsm) file for inspection and validation.

Prints cell text, sheet names, named ranges, and comments to stdout,
one item per line, so the output can be piped to grep, awk, or other tools.

Usage:
    python extract.py <file.xlsx>

Examples:
    python extract.py workbook.xlsx
    python extract.py edited.xlsx | grep -iE "xxxx|lorem|ipsum"
    python extract.py workbook.xlsx | grep "Sheet1"
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def extract(input_file: str) -> str:
    src = Path(input_file)
    if not src.exists():
        return f"Error: {input_file} does not exist"
    if src.suffix.lower() not in (".xlsx", ".xlsm"):
        return f"Error: {input_file} must be a .xlsx or .xlsm file"

    try:
        wb = openpyxl.load_workbook(str(src), data_only=True)
        lines = []

        # Sheet names
        lines.append("<!-- sheet names -->")
        for name in wb.sheetnames:
            lines.append(f"[sheet] {name}")

        # Cell text per sheet
        for sheet in wb.worksheets:
            lines.append(f"<!-- sheet: {sheet.title} -->")
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        text = cell.value.strip()
                        if text:
                            lines.append(f"{sheet.title}!{cell.coordinate}: {text}")

            # Cell comments/notes
            if sheet.comments:
                for coord, comment in sheet.comments.items():
                    raw = comment.text
                    text = raw.strip() if isinstance(raw, str) else ""
                    if text:
                        lines.append(f"[comment] {sheet.title}!{coord}: {text}")

        # Named ranges (defined names)
        if wb.defined_names:
            defns = list(wb.defined_names)
            if defns:
                lines.append("<!-- named ranges -->")
                for dn in defns:
                    lines.append(f"[named range] {dn.name}: {dn.value}")

        return "\n".join(lines)
    except Exception as e:
        return f"Error: {e}"


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract text from an Excel file")
    parser.add_argument("input_file", help="Excel file to read (.xlsx or .xlsm)")
    args = parser.parse_args()

    result = extract(args.input_file)
    print(result)
    if result.startswith("Error"):
        sys.exit(1)
